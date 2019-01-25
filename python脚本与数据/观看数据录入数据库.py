#! python3
# 把excel数据录入数据库

import openpyxl, pymysql

# 打开excel
infoExcel = openpyxl.load_workbook('观看数据.xlsx')
infoSheetNames = infoExcel.sheetnames
infoSheet = infoExcel[infoSheetNames[0]]
totalRow = infoSheet.max_row

# 打开数据库连接
db = pymysql.connect("localhost", "pilipala6868", "hz19971023", "movielist" )
cursor = db.cursor()

failID = []
# 将excel信息逐条录入
for i in range(totalRow):
	# sql语句
	tempID = i + 1
	rowContent = [t.value for t in infoSheet['A%d'%tempID : 'J%d'%tempID][0]]
	rowContent[0] = str(rowContent[0]).split(' ')[0]  # 去掉日期后具体到的时间

	sqlConmend = """insert into record values (%s, %s, "%s", "宿舍", "%s", %s)""" % (str(tempID), str(tempID), rowContent[0], rowContent[1], rowContent[2])

	# print(sqlConmend)

	# 执行SQL
	try:
		cursor.execute(sqlConmend)
		db.commit()
		print("Infomation %d insert success." % (i+1))
	except:
		print("Error at No.%d" % (i+1))
		failID.append(i+1)

	# break

# 关闭数据库连接
db.close()

# 报结果
if len(failID) == 0:
	print("ALL SUCCESS!")
else:
	print("FailID:")
	for i in failID:
		print(str(i) + ' ', end='')
