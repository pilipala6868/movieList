#! python3
# 把excel数据录入数据库

import openpyxl, pymysql

# 打开excel
infoExcel = openpyxl.load_workbook('演员数据.xlsx')
infoSheetNames = infoExcel.sheetnames
infoSheet = infoExcel[infoSheetNames[0]]
totalRow = infoSheet.max_row

# 打开数据库连接
db = pymysql.connect("localhost", "pilipala6868", "hz19971023", "movielist" )
cursor = db.cursor()

failID = []
# 将excel信息逐条录入
for i in range(totalRow):
	# sql语句
	tempID = i + 1
	rowContent = [t.value for t in infoSheet['A%d'%tempID : 'E%d'%tempID][0]]

	sqlConmend = """insert into actor values (%s, "%s", "%s", "%s", "%s")""" % (rowContent[0], rowContent[1], rowContent[2], rowContent[3], rowContent[4])

	# print(sqlConmend)

	# 执行SQL
	try:
		cursor.execute(sqlConmend)
		db.commit()
		print("Infomation %d insert success." % tempID)
	except:
		print("Error at No.%d" % tempID)
		failID.append(tempID)

	# break

# 关闭数据库连接
db.close()

# 报结果
if len(failID) == 0:
	print("ALL SUCCESS!")
else:
	print("FailID:")
	for i in failID:
		print(str(i) + ' ', end='')
