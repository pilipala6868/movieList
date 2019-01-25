#! python3
# 把excel数据录入数据库

import openpyxl, pymysql

# 打开excel
infoExcel = openpyxl.load_workbook('电影数据.xlsx')
infoSheetNames = infoExcel.sheetnames
infoSheet = infoExcel[infoSheetNames[0]]
totalRow = infoSheet.max_row

# 打开数据库连接
db = pymysql.connect("localhost", "pilipala6868", "hz19971023", "movielist" )
cursor = db.cursor()

failID = []
# 将excel信息逐条录入
for i in range(totalRow):
	# sql语句
	tempID = i + 1
	rowContent = [t.value for t in infoSheet['A%d'%tempID : 'J%d'%tempID][0]]
	if rowContent[8] == None:
		rowContent[8] = 'NULL'
	if rowContent[9] == None:
		rowContent[9] = 'NULL'

	sqlConmend = """insert into movie values (%s, "%s", "%s", "%s", "%s", "%s", "%s", %s, %s, %s, %s, "%s")""" % (str(tempID), rowContent[0], rowContent[1], rowContent[2], rowContent[3], rowContent[4], rowContent[5], rowContent[6], rowContent[7], rowContent[8], rowContent[9], '%d.jpg' % tempID)

	# print(sqlConmend)

	# 执行SQL
	try:
		cursor.execute(sqlConmend)
		db.commit()
		print("Infomation %d insert success." % (i+1))
	except:
		print("Error at No.%d" % (i+1))
		failID.append(i+1)

# # 把空的英文名改为NULL
# try:
# 	cursor.execute("""update movie set name_en = NULL where name_en = 'None'""")
# except:
# 	print("Error at updating NULL.")

# 关闭数据库连接
db.close()

# 报结果
if len(failID) == 0:
	print("ALL SUCCESS!")
else:
	print("FailID:")
	for i in failID:
		print(str(i) + ' ', end='')
